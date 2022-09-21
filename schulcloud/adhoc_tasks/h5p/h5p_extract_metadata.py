import sys
from typing import List, Union, IO, Optional, Literal
import re
import openpyxl
from openpyxl.utils import get_column_letter


class Metadata:
    def __init__(self, title: str, publisher: str, keywords: List[str], order: str,
                 permission: List[Literal['ALLE', 'NDS', 'BRB', 'THR']],
                 collection: Optional['Collection'] = None, licence: Optional[str] = None):
        self.title = title
        self.publisher = publisher
        self.keywords = keywords
        self.order = order
        self.license = licence
        self.permission = permission
        self.collection = collection
        if collection is not None and type(collection) != str:
            self.collection.add_child(self)


class Collection:
    def __init__(self, name: str):
        self.name = name
        self.children: List[Metadata] = []

    def add_child(self, child: Metadata):
        self.children.append(child)


class MetadataFile:
    class COLUMN:
        COLLECTION = 1
        TITLE_AND_ORDER = 2
        ORDER = 3
        TITLE = 4
        FILENAME = 5
        KEYWORDS = 6
        PUBLISHER = 7
        LICENSE = 8
        PERMISSION = 9
        COUNT = 9

    def __init__(self, file: Union[str, IO]):
        self.file = file
        self.collections: List[Collection] = []
        self.single_files: List[Metadata] = []
        self.workbook = openpyxl.load_workbook(filename=file, data_only=True)
        self.o_sheet = self.workbook["Tabelle1"]
        self.validate_sheet()
        self.parse()

    def close(self):
        self.workbook.close()
        if not isinstance(self.file, str):
            self.file.close()

    def validate_sheet(self):
        # Check required fields
        required_fields = [self.COLUMN.TITLE, self.COLUMN.FILENAME, self.COLUMN.PERMISSION]
        for row in range(1, self.o_sheet.max_row + 1):
            for column in range(1, self.COLUMN.COUNT + 1):
                if column in required_fields:
                    if not self.o_sheet.cell(row=row, column=column).value:
                        raise ParsingError(f'Empty required cell: column: {get_column_letter(column)} row: {row} at '
                                           f'{self.file.name}')

        # Check for collection
        if self.o_sheet.cell(row=1, column=self.COLUMN.COLLECTION).value:
            collection = self.o_sheet.cell(row=1, column=self.COLUMN.COLLECTION).value
            for row in range(1, self.o_sheet.max_row + 1):
                other_collection = self.o_sheet.cell(row=row, column=self.COLUMN.COLLECTION).value
                if not other_collection == collection:
                    raise ParsingError(f'Multiple collection or spelling mistake in row {row} at {self.file.name}.')
        # Check permissions of collection
            permissions = []
            for row in range(1, self.o_sheet.max_row + 1):
                permissions_raw = self.o_sheet.cell(row=row, column=self.COLUMN.PERMISSION).value
                permissions.append(permissions_raw)
            first_permission = permissions[0]
            for permission in permissions:
                if permission not in ('NDS', 'BRB', 'THR', 'ALLE'):
                    raise ParsingError(f'Spelling mistake or unknown permission: {permission} at {self.file.name}')
                elif permission != first_permission:
                    raise ParsingError(f'Permissions in the collection: "{collection}" of "{self.file.name}" '
                                       f'are not matching! ({first_permission} != {permission})')

    def parse(self):
        for row in range(1, self.o_sheet.max_row + 1):
            order = self.o_sheet.cell(row=row, column=self.COLUMN.ORDER).value
            prefix = self.fill_zeros(str(order)) + ' ' if order else ""
            title = prefix + str(self.o_sheet.cell(row=row, column=self.COLUMN.TITLE).value)
            keywords_raw = self.o_sheet.cell(row=row, column=self.COLUMN.KEYWORDS).value
            keywords = re.findall(r'\w+', keywords_raw)
            publisher = self.o_sheet.cell(row=row, column=self.COLUMN.PUBLISHER).value
            licence = self.o_sheet.cell(row=row, column=self.COLUMN.LICENSE).value
            permissions = re.findall(r'\w+', self.o_sheet.cell(row=row, column=self.COLUMN.PERMISSION).value)

            collection_name = self.o_sheet.cell(row=row, column=self.COLUMN.COLLECTION).value
            if collection_name:
                for c in self.collections:
                    if c.name == collection_name:
                        collection = c
                        break
                else:
                    collection = Collection(collection_name)
                    self.collections.append(collection)
                Metadata(title, publisher, keywords, order, permissions, collection, licence)
            else:
                self.single_files.append(Metadata(title, publisher, keywords, order, permissions, licence=licence))

    def check_for_files(self, filenames: List[str]):
        existing_filenames = []
        missing_filenames = []
        for row in range(1, self.o_sheet.max_row + 1):
            existing_filenames.append(self.o_sheet.cell(row=row, column=self.COLUMN.FILENAME).value)
        for filename in filenames:
            if filename not in existing_filenames:
                missing_filenames.append(filename)
        if missing_filenames:
            print('Missing in excel file:', file=sys.stderr)
            for filename in missing_filenames:
                print(f'  {filename}', file=sys.stderr)
            raise ParsingError('The excel file is missing metadata')

    def get_collection(self):
        return self.o_sheet.cell(row=1, column=self.COLUMN.COLLECTION).value

    def get_collection_permission(self):
        return self.o_sheet.cell(row=1, column=self.COLUMN.PERMISSION).value

    def get_keywords(self):
        keywords_raw = self.o_sheet.cell(row=1, column=self.COLUMN.KEYWORDS).value
        keywords = re.findall(r'\w+', keywords_raw)
        return keywords

    def get_publisher(self):
        return self.o_sheet.cell(row=1, column=self.COLUMN.PUBLISHER).value

    def get_license(self):
        licence = self.o_sheet.cell(row=1, column=self.COLUMN.LICENSE).value
        if licence is not None:
            licence = licence.rsplit('(', 1)
            licence = licence[0]
        return licence

    def find_metadata_by_file_name(self, h5p_file: str):
        result = []
        # looking for exact match
        for row in range(1, self.o_sheet.max_row + 1):
            if self.o_sheet.cell(row=row, column=self.COLUMN.FILENAME).value == h5p_file:
                result.append(row)
        if not result:
            # looking for rough match (like relative paths etc.)
            for row in range(1, self.o_sheet.max_row + 1):
                if h5p_file in self.o_sheet.cell(row=row, column=self.COLUMN.FILENAME).value:
                    result.append(row)
            if not result:
                raise RuntimeError(f'No metadata found for {h5p_file}')

        if len(result) == 1:
            return result[0]
        elif len(result) > 1:
            raise RuntimeError(f'Multiple metadata matches for {h5p_file}')

    def get_metadata(self, h5p_file: str):
        row = self.find_metadata_by_file_name(h5p_file)

        collection = self.o_sheet.cell(row=row, column=self.COLUMN.COLLECTION).value
        order = self.o_sheet.cell(row=row, column=self.COLUMN.ORDER).value
        prefix = self.fill_zeros(str(order)) + ' ' if order else ""
        title = prefix + str(self.o_sheet.cell(row=row, column=self.COLUMN.TITLE).value)
        keywords_raw = self.o_sheet.cell(row=row, column=self.COLUMN.KEYWORDS).value
        keywords = re.findall(r'\w+', keywords_raw)
        publisher = self.o_sheet.cell(row=row, column=self.COLUMN.PUBLISHER).value
        licence = self.o_sheet.cell(row=row, column=self.COLUMN.LICENSE).value
        if licence is not None:
            licence = licence.rsplit('(', 1)
            licence = licence[0]
        permission = re.findall(r'\w+', self.o_sheet.cell(row=row, column=self.COLUMN.PERMISSION).value)

        return Metadata(title, publisher, keywords, order, permission, collection, licence)

    def fill_zeros(self, order: str):
        max_length = len(str(self.o_sheet.max_row))
        zero = (max_length - len(order)) * '0'
        return zero + order


class ParsingError(Exception):
    def __init__(self, msg: str):
        super(ParsingError, self).__init__(msg)
