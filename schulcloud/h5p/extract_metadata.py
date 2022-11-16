import sys
from typing import List, Union, IO, Optional, Set
import re

import openpyxl
from openpyxl.utils import get_column_letter


class Metadata:
    def __init__(self, filepath: str, title: str, publisher: str, keywords: List[str], order: str,
                 permission: List[str],
                 collection: Optional['Collection'] = None, license: Optional[str] = None):
        self.filepath = filepath
        self.title = title
        self.publisher = publisher
        self.keywords = keywords
        self.order = order
        self.license = license
        self.permission = permission
        self.collection = collection
        if collection is not None and type(collection) != str:
            self.collection.add_child(self)


class Collection:
    def __init__(self, name: str):
        self.name = name
        self.publishers: Set[str] = set()
        self.keywords: Set[str] = set()
        self.licenses: Set[str] = set()
        self.permissions: Set[str] = set()
        self.children: List[Metadata] = []

    def add_child(self, child: Metadata):
        self.children.append(child)
        self.publishers.add(child.publisher)
        self.keywords.update(child.keywords)
        self.licenses.add(child.license)
        self.permissions.update(child.permission)


class MetadataFile:
    class COLUMN:
        COLLECTION = 1
        TITLE_AND_ORDER = 2
        ORDER = 3
        TITLE = 4
        FILENAME = 5
        KEYWORDS = 6
        PUBLISHER = 7
        LICENSE = 8
        PERMISSION = 9
        COUNT = 9

    def __init__(self, file: Union[str, IO]):
        self.file = file
        self.collections: List[Collection] = []  # collections of files
        self.single_files: List[Metadata] = []  # files outside any collection
        self.workbook = openpyxl.load_workbook(filename=file, data_only=True)
        self.o_sheet = self.workbook["Tabelle1"]
        self._validate_sheet()
        self._parse()

    def close(self):
        self.workbook.close()
        if not isinstance(self.file, str):
            self.file.close()

    def _validate_sheet(self):
        # Check required fields
        required_fields = [self.COLUMN.TITLE, self.COLUMN.FILENAME, self.COLUMN.PERMISSION]
        for row in range(1, self.o_sheet.max_row + 1):
            for column in range(1, self.COLUMN.COUNT + 1):
                if column in required_fields:
                    if not self.o_sheet.cell(row=row, column=column).value:
                        raise ParsingError(f'Empty required cell: column: {get_column_letter(column)} row: {row} at '
                                           f'{self.file.name}')

        # Check for collection
        if self.o_sheet.cell(row=1, column=self.COLUMN.COLLECTION).value:
            collection = self.o_sheet.cell(row=1, column=self.COLUMN.COLLECTION).value
            for row in range(1, self.o_sheet.max_row + 1):
                other_collection = self.o_sheet.cell(row=row, column=self.COLUMN.COLLECTION).value
                if not other_collection == collection:
                    raise ParsingError(f'Multiple collection or spelling mistake in row {row} at {self.file.name}.')
        # Check permissions of collection
            permissions = []
            for row in range(1, self.o_sheet.max_row + 1):
                permissions_raw = self.o_sheet.cell(row=row, column=self.COLUMN.PERMISSION).value
                permissions.append(permissions_raw)
            first_permission = permissions[0]
            for permission in permissions:
                if permission not in ('NDS', 'BRB', 'THR', 'ALLE'):
                    raise ParsingError(f'Spelling mistake or unknown permission: {permission} at {self.file.name}')
                elif permission != first_permission:
                    raise ParsingError(f'Permissions in the collection: "{collection}" of "{self.file.name}" '
                                       f'are not matching! ({first_permission} != {permission})')

    def _parse(self):
        for row in range(1, self.o_sheet.max_row + 1):
            order_raw: Optional[Union[int, str]] = self.o_sheet.cell(row=row, column=self.COLUMN.ORDER).value
            order = str(order_raw) if order_raw else ''
            prefix: str = self._fill_zeros(order) + '. ' if order else ''
            title: str = prefix + str(self.o_sheet.cell(row=row, column=self.COLUMN.TITLE).value)
            filepath: str = self.o_sheet.cell(row=row, column=self.COLUMN.FILENAME).value
            keywords_raw: str = self.o_sheet.cell(row=row, column=self.COLUMN.KEYWORDS).value
            keywords: List[str] = re.findall(r'\w+', keywords_raw)
            publisher: str = self.o_sheet.cell(row=row, column=self.COLUMN.PUBLISHER).value
            license_raw: str = self.o_sheet.cell(row=row, column=self.COLUMN.LICENSE).value or ''
            license = license_raw.rsplit('(', 1)[0]
            permissions: List[str] = re.findall(r'\w+', self.o_sheet.cell(row=row, column=self.COLUMN.PERMISSION).value)

            collection_name: Optional[str] = self.o_sheet.cell(row=row, column=self.COLUMN.COLLECTION).value
            if collection_name:
                for c in self.collections:
                    if c.name == collection_name:
                        collection = c
                        break
                else:
                    collection = Collection(collection_name)
                    self.collections.append(collection)
                Metadata(filepath, title, publisher, keywords, order, permissions, collection, license)
            else:
                self.single_files.append(Metadata(filepath, title, publisher, keywords, order, permissions, license=license))

    def _fill_zeros(self, order: str):
        max_length = len(str(self.o_sheet.max_row))
        zero = (max_length - len(order)) * '0'
        return zero + order

    def check_for_files(self, filenames: List[str]):
        existing_filenames = []
        missing_filenames = []
        for row in range(1, self.o_sheet.max_row + 1):
            existing_filenames.append(self.o_sheet.cell(row=row, column=self.COLUMN.FILENAME).value)
        for filename in filenames:
            if filename not in existing_filenames:
                missing_filenames.append(filename)
        if missing_filenames:
            print('Missing in excel file:', file=sys.stderr)
            for filename in missing_filenames:
                print(f'  {filename}', file=sys.stderr)
            raise ParsingError('The excel file is missing metadata')


class ParsingError(Exception):
    def __init__(self, msg: str):
        super(ParsingError, self).__init__(msg)
